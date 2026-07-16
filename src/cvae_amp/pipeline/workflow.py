"""Full AMP design pipeline: generate → dedup → filter → predict.

Supports all three generative models (VAE, CVAE, CVAE+Pred).
"""

from __future__ import annotations

import os
import sys
import time
import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
import torch

from cvae_amp.config.paths import (
    ROOT, DATA_RAW, MODELS_GEN, MODELS_PRED, RESULTS_PIPELINE, ensure_dirs,
)
from cvae_amp.config.defaults import (
    DEFAULT_TEMPERATURE, DEFAULT_TOP_K, DEFAULT_MIN_LEN, DEFAULT_MAX_LEN,
    PIPELINE_NUM_GEN, PIPELINE_CDHIT_IDENTITY, PIPELINE_BLAST_EVALUE,
    PIPELINE_BLAST_IDENTITY_CUTOFF, PIPELINE_PRED_THRESHOLD,
)
from cvae_amp.generation.models import VAE, CVAE, CVAE_Pred, DEVICE
from cvae_amp.generation.sampling import sample_vae, sample_cvae
from cvae_amp.pipeline.filters import valid_sequence
from cvae_amp.pipeline.dedup import run_cdhit
from cvae_amp.pipeline.homology import build_blast_db, run_blastp, parse_blast_hits


MODEL_LOADERS = {
    "vae": (VAE, "vae_model.pth"),
    "cvae": (CVAE, "cvae_model.pth"),
    "cvae_pred": (CVAE_Pred, "cvae_pred_model.pth"),
}


class AMPWorkflow:
    def __init__(
        self,
        model_name: str = "cvae_pred",
        target: tuple[float, float, float] = (1.0, 1.0, 1.0),
        num_gen: int = PIPELINE_NUM_GEN,
        temperature: float = DEFAULT_TEMPERATURE,
        top_k: int = DEFAULT_TOP_K,
        min_len: int = DEFAULT_MIN_LEN,
        max_len: int = DEFAULT_MAX_LEN,
        work_dir: Path | None = None,
        train_xlsx: Path | None = None,
        vae_seq_file: Path | None = None,
        cdhit_bin: str | None = None,
        blast_bin: str | None = None,
    ) -> None:
        self.model_name = model_name
        self.target = list(target)
        self.num_gen = num_gen
        self.temperature = temperature
        self.top_k = top_k
        self.min_len = min_len
        self.max_len = max_len

        self.work_dir = work_dir or RESULTS_PIPELINE
        self.work_dir.mkdir(parents=True, exist_ok=True)

        self.train_xlsx = train_xlsx or DATA_RAW / "VAE_train_real.xlsx"
        self.vae_seq_file = vae_seq_file or ROOT / "vae_seq.txt"
        self.cdhit_bin = cdhit_bin
        self.blast_bin = blast_bin

        self._model = None

    def _load_model(self):
        model_cls, ckpt_name = MODEL_LOADERS[self.model_name]
        ckpt_path = MODELS_GEN / ckpt_name
        model = model_cls().to(DEVICE)
        model.load_state_dict(torch.load(str(ckpt_path), map_location=DEVICE))
        model.eval()
        n_params = sum(p.numel() for p in model.parameters())
        print(f"  Model loaded: {self.model_name} ({n_params:,} params)")
        return model

    def step_generate(self) -> Path:
        print("=" * 60)
        print(f"Step 1: Generating {self.num_gen:,} peptides with {self.model_name.upper()}")
        print("=" * 60)

        model = self._load_model()
        is_vae = (self.model_name == "vae")

        unique: set[str] = set()
        attempts = 0
        max_attempts = self.num_gen * 20
        batch_size = 64
        start = time.time()

        while len(unique) < self.num_gen and attempts < max_attempts:
            n_batch = min(batch_size, (self.num_gen - len(unique)) * 2)
            if is_vae:
                peptides = sample_vae(model, temperature=self.temperature,
                                      top_k=self.top_k, min_len=self.min_len,
                                      num_samples=n_batch)
            else:
                peptides = sample_cvae(model, self.target, temperature=self.temperature,
                                       top_k=self.top_k, min_len=self.min_len,
                                       num_samples=n_batch)
            for p in peptides:
                if valid_sequence(p, self.min_len, self.max_len):
                    unique.add(p)
                    if len(unique) >= self.num_gen:
                        break
            attempts += n_batch

        peptides = list(unique)[:self.num_gen]
        elapsed = time.time() - start
        print(f"  Done: {len(peptides):,} peptides in {elapsed:.0f}s")

        out_xlsx = self.work_dir / "generated_30k.xlsx"
        pd.DataFrame({"sequence": peptides, "length": [len(p) for p in peptides]}).to_excel(out_xlsx, index=False)

        fasta_path = self.work_dir / "generated_30k.fasta"
        with open(fasta_path, "w") as f:
            for i, s in enumerate(peptides):
                f.write(f">gen{i}\n{s}\n")

        print(f"  Saved: {out_xlsx}, {fasta_path}\n")
        return fasta_path

    def step_prepare_train(self) -> Path:
        print("=" * 60)
        print("Step 2: Preparing training set FASTA")
        print("=" * 60)

        wb = openpyxl.load_workbook(self.train_xlsx)
        ws = wb.active
        seqs = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)
                if row[1] and len(str(row[1])) <= 50]
        seqs = list(dict.fromkeys(seqs))

        fasta_path = self.work_dir / "train_all.fasta"
        with open(fasta_path, "w") as f:
            for i, s in enumerate(seqs):
                f.write(f">train{i}\n{s}\n")

        print(f"  {len(seqs)} unique training sequences\n")
        return fasta_path

    def step_cdhit(self, generated_fasta: Path) -> Path:
        print("=" * 60)
        print("Step 3: CD-HIT deduplication")
        print("=" * 60)
        output = self.work_dir / "candidates_cdhit70"
        return run_cdhit(generated_fasta, output, identity=PIPELINE_CDHIT_IDENTITY, cdhit_bin=self.cdhit_bin)

    def step_blast(self, cdhit_fasta: Path, train_fasta: Path) -> Path:
        print("=" * 60)
        print("Step 4: BLASTP against training set")
        print("=" * 60)
        db_name = str(self.work_dir / "train_db")
        build_blast_db(train_fasta, db_name, blast_bin=self.blast_bin)
        output = self.work_dir / "blast_results.txt"
        return run_blastp(cdhit_fasta, db_name, output, evalue=PIPELINE_BLAST_EVALUE, blast_bin=self.blast_bin)

    def step_filter(self, cdhit_fasta: Path, blast_out: Path) -> Path:
        print("=" * 60)
        print("Step 5: Homology filter + manual sequences")
        print("=" * 60)

        blast_hits = parse_blast_hits(blast_out, PIPELINE_BLAST_IDENTITY_CUTOFF)
        print(f"  Sequences with homologs: {len(blast_hits)}")

        all_seqs: list[str] = []
        with open(cdhit_fasta) as f:
            current = ""
            for line in f:
                if line.startswith(">"):
                    if current:
                        all_seqs.append(current)
                    current = ""
                else:
                    current += line.strip()
            if current:
                all_seqs.append(current)

        kept = [s for i, s in enumerate(all_seqs)
                if f"gen{i}" not in blast_hits and f"seq{i}" not in blast_hits]
        print(f"  Passed all filters: {len(kept)}")

        manual_seqs: list[str] = []
        if self.vae_seq_file.exists():
            with open(self.vae_seq_file) as f:
                manual_seqs = [line.strip() for line in f if line.strip()]
            print(f"  Manual sequences: {len(manual_seqs)}")

        final = manual_seqs + [s for s in kept if s not in manual_seqs]
        final = list(dict.fromkeys(final))
        print(f"  Final: {len(final)} sequences")

        out_path = self.work_dir / "candidates_final_filtered.xlsx"
        pd.DataFrame({"sequence": final, "length": [len(s) for s in final]}).to_excel(out_path, index=False)
        print(f"  Saved: {out_path}\n")
        return out_path

    def step_predict(self, candidates_xlsx: Path) -> Path:
        print("=" * 60)
        print("Step 6: Activity prediction (AMP / AEP / HP)")
        print("=" * 60)

        from cvae_amp.prediction.features import FeatureEncoder
        from cvae_amp.prediction.loaders import load_amp, load_aep, load_hp

        device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        print("  Loading models...")
        model_amp = load_amp(device)
        model_aep = load_aep()
        model_hp = load_hp()

        feature_dir = ROOT / "data" / "features"
        print("  Encoding features...")
        enc_amp = FeatureEncoder(str(feature_dir / "20aa_pc7_feature.xlsx"))
        enc_hp = FeatureEncoder(str(feature_dir / "20aa_pc5_1_feature.xlsx"))

        feat_amp = enc_amp.encode_file(str(candidates_xlsx))
        feat_hp = enc_hp.encode_file(str(candidates_xlsx))
        n = len(feat_amp)
        print(f"  Predicting {n} sequences...")

        # AMP
        t = torch.tensor(feat_amp, dtype=torch.float32).to(device)
        with torch.no_grad():
            amp_pred = model_amp(t).cpu().numpy().flatten()

        # AEP
        aep_pred = model_aep.predict_proba(feat_amp.reshape(n, -1))[:, 1]

        # HP
        hp_pred = model_hp.predict_proba(feat_hp.reshape(n, -1))[:, 1]

        # Build output
        print("  Writing results...")
        wb_in = openpyxl.load_workbook(candidates_xlsx, data_only=True)
        ws_in = wb_in.active

        wb = openpyxl.Workbook()
        ws_all = wb.active
        ws_all.title = "all_predictions"
        hdrs = ["sequence", "length", "amp_Pred", "aep_Pred", "hp_Pred", "pass_all"]
        ws_all.append(hdrs)
        high_conf = []

        for i in range(2, ws_in.max_row + 1):
            idx = i - 2
            seq = ws_in.cell(row=i, column=1).value or ""
            length = ws_in.cell(row=i, column=2).value or len(str(seq))
            amp = float(amp_pred[idx])
            aep = float(aep_pred[idx])
            hp = float(hp_pred[idx])
            passed = "Y" if (amp > PIPELINE_PRED_THRESHOLD and aep > PIPELINE_PRED_THRESHOLD
                             and hp > PIPELINE_PRED_THRESHOLD) else "N"
            row_data = (seq, length, amp, aep, hp, passed)
            ws_all.append(row_data)
            if passed == "Y":
                high_conf.append(row_data)

        for col_letter, width in [("A", 40), ("B", 8), ("C", 12), ("D", 12), ("E", 12), ("F", 10)]:
            ws_all.column_dimensions[col_letter].width = width

        ws_hc = wb.create_sheet("high_confidence")
        ws_hc.append(hdrs)
        for row_data in high_conf:
            ws_hc.append(row_data)

        out_path = self.work_dir / "pipeline_results.xlsx"
        wb.save(out_path)
        print(f"  Done -> {out_path}")
        print(f"    all_predictions: {n} | high_confidence: {len(high_conf)}")
        print(f"    AMP mean={amp_pred.mean():.4f}  AEP mean={aep_pred.mean():.4f}  HP mean={hp_pred.mean():.4f}")
        return out_path

    def run(self) -> Path:
        ensure_dirs()
        t0 = time.time()

        generated_fasta = self.step_generate()
        train_fasta = self.step_prepare_train()
        cdhit_fasta = self.step_cdhit(generated_fasta)
        blast_out = self.step_blast(cdhit_fasta, train_fasta)
        candidates_xlsx = self.step_filter(cdhit_fasta, blast_out)
        results_xlsx = self.step_predict(candidates_xlsx)

        elapsed = time.time() - t0
        print(f"\n{'=' * 60}")
        print(f"Pipeline complete in {elapsed:.0f}s")
        print(f"Final output: {results_xlsx}")
        print(f"{'=' * 60}")
        return results_xlsx


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Full CVAE → predict pipeline")
    parser.add_argument("--model", default="cvae_pred", choices=["vae", "cvae", "cvae_pred"])
    parser.add_argument("--target", type=str, default="1.0,1.0,1.0")
    parser.add_argument("--num", type=int, default=1000)
    parser.add_argument("--temperature", type=float, default=0.7)
    parser.add_argument("--work-dir", type=str, default=None)
    args = parser.parse_args()

    target = tuple(float(x) for x in args.target.split(","))
    work_dir = Path(args.work_dir) if args.work_dir else None

    wf = AMPWorkflow(
        model_name=args.model,
        target=target,
        num_gen=args.num,
        temperature=args.temperature,
        work_dir=work_dir,
    )
    wf.run()
