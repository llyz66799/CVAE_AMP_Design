"""Train BiLSTM+Attention model with Optuna hyperparameter optimization.

Refactored from attention_v3.py (folder 33). Uses LayerNorm, 2-layer LSTM,
AdamW optimizer, CosineAnnealingLR scheduler, and AUC-based early stopping.
"""

from __future__ import annotations

import os
import time
import random
import argparse

import numpy as np
import torch
import torch.nn as nn
from torch.utils.data import TensorDataset, DataLoader
from torch.optim.lr_scheduler import CosineAnnealingLR
from sklearn.metrics import accuracy_score, precision_score, recall_score, roc_auc_score
import optuna
import openpyxl as xl

from cvae_amp.config.paths import DATA_RAW, DATA_FEATURES, MODELS_PRED, RESULTS
from cvae_amp.prediction.models import AMPAttentionModel
from cvae_amp.prediction.features import FeatureEncoder


def set_seed(seed: int = 16) -> None:
    np.random.seed(seed)
    random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)
        torch.backends.cudnn.deterministic = True
        torch.backends.cudnn.benchmark = False


def evaluate_model(model: nn.Module, loader: DataLoader, device: torch.device) -> tuple[float, float, float, float]:
    model.eval()
    all_preds, all_targets = [], []
    with torch.no_grad():
        for batch_x, batch_y in loader:
            batch_x, batch_y = batch_x.to(device), batch_y.to(device)
            preds = model(batch_x)
            all_preds.extend(preds.cpu().numpy())
            all_targets.extend(batch_y.cpu().numpy())

    all_preds = np.array(all_preds)
    all_targets = np.array(all_targets)
    pred_classes = (all_preds > 0.5).astype(int)

    acc = accuracy_score(all_targets, pred_classes)
    pre = precision_score(all_targets, pred_classes, zero_division=0)
    rec = recall_score(all_targets, pred_classes, zero_division=0)
    try:
        auc = roc_auc_score(all_targets, all_preds)
    except ValueError:
        auc = 0.5
    return acc, pre, rec, auc


def load_data(
    task: str,
    feature_name: str,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, int]:
    task_files = {
        "amp": ("amp_train6450.xlsx", "amp_val2149.xlsx", "amp_test2149.xlsx"),
        "aep": ("aep_train1215.xlsx", "aep_val404.xlsx", "aep_test404.xlsx"),
        "hp": ("hp_train1250.xlsx", "hp_val416.xlsx", "hp_test416.xlsx"),
    }
    train_f, val_f, test_f = task_files[task]

    encoder = FeatureEncoder(str(DATA_FEATURES / f"20aa_{feature_name}_feature.xlsx"))

    x_train = encoder.encode_file(str(DATA_RAW / train_f))
    y_train = _read_labels(DATA_RAW / train_f)
    x_val = encoder.encode_file(str(DATA_RAW / val_f))
    y_val = _read_labels(DATA_RAW / val_f)
    x_test = encoder.encode_file(str(DATA_RAW / test_f))
    y_test = _read_labels(DATA_RAW / test_f)

    return x_train, y_train, x_val, y_val, x_test, y_test, encoder.dim


def _read_labels(path) -> np.ndarray:
    wb = xl.load_workbook(str(path))
    ws = wb.active
    labels = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        labels.append([int(row[-1])])
    return np.array(labels, dtype=np.float32)


def make_loader(x: np.ndarray, y: np.ndarray, batch_size: int = 128, shuffle: bool = False) -> DataLoader:
    xt = torch.tensor(x, dtype=torch.float32)
    yt = torch.tensor(y, dtype=torch.float32)
    return DataLoader(TensorDataset(xt, yt), batch_size=batch_size, shuffle=shuffle)


class Objective:
    def __init__(self, train_loader: DataLoader, val_loader: DataLoader, input_dim: int, device: torch.device):
        self.train_loader = train_loader
        self.val_loader = val_loader
        self.input_dim = input_dim
        self.device = device

    def __call__(self, trial: optuna.Trial) -> float:
        lr = trial.suggest_float("lr", 1e-4, 5e-3, log=True)
        dropout = trial.suggest_float("dropout", 0.1, 0.5)
        lstm_hidden = trial.suggest_categorical("lstm_hidden", [32, 64, 128])
        weight_decay = trial.suggest_float("weight_decay", 1e-6, 1e-3, log=True)

        set_seed(16)
        model = AMPAttentionModel(
            input_dim=self.input_dim,
            lstm_hidden=lstm_hidden,
            num_layers=2,
            dropout_rate=dropout,
        ).to(self.device)

        optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=weight_decay)
        scheduler = CosineAnnealingLR(optimizer, T_max=300, eta_min=1e-6)
        criterion = nn.BCELoss()

        best_val_auc = 0.0
        patience = 100
        counter = 0

        for epoch in range(300):
            model.train()
            for batch_x, batch_y in self.train_loader:
                batch_x, batch_y = batch_x.to(self.device), batch_y.to(self.device)
                optimizer.zero_grad()
                loss = criterion(model(batch_x), batch_y)
                loss.backward()
                optimizer.step()
            scheduler.step()

            _, _, _, val_auc = evaluate_model(model, self.val_loader, self.device)

            if val_auc > best_val_auc:
                best_val_auc = val_auc
                counter = 0
                trial.set_user_attr("best_model_state", model.state_dict())
                trial.set_user_attr("best_epoch", epoch + 1)
            else:
                counter += 1

            if counter >= patience:
                break

        return best_val_auc


def train_attention(
    task: str = "amp",
    feature: str = "pc7",
    n_trials: int = 30,
) -> str:
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Using device: {device}")
    print(f"Task: {task}, Feature: {feature}, Trials: {n_trials}")

    x_train, y_train, x_val, y_val, x_test, y_test, input_dim = load_data(task, feature)

    train_loader = make_loader(x_train, y_train, batch_size=128, shuffle=True)
    val_loader = make_loader(x_val, y_val, batch_size=256)
    test_loader = make_loader(x_test, y_test, batch_size=256)

    objective = Objective(train_loader, val_loader, input_dim, device)
    study = optuna.create_study(direction="maximize")
    study.optimize(objective, n_trials=n_trials)

    best_trial = study.best_trial
    print(f"Best val AUC: {best_trial.value:.4f}")
    print(f"Best params: {best_trial.params}")

    # Build final model with best params and load best epoch weights
    final_model = AMPAttentionModel(
        input_dim=input_dim,
        lstm_hidden=best_trial.params["lstm_hidden"],
        num_layers=2,
        dropout_rate=best_trial.params["dropout"],
    ).to(device)
    final_model.load_state_dict(best_trial.user_attrs["best_model_state"])

    t_acc, t_pre, t_rec, t_auc = evaluate_model(final_model, test_loader, device)
    print(f"Test: Acc={t_acc:.4f}, Pre={t_pre:.4f}, Rec={t_rec:.4f}, AUC={t_auc:.4f}")

    os.makedirs(str(MODELS_PRED), exist_ok=True)
    save_path = MODELS_PRED / f"BestModel_{task.upper()}_{feature}.pth"
    torch.save(best_trial.user_attrs["best_model_state"], str(save_path))
    print(f"Model saved to {save_path}")

    # Save optimization results
    results_path = str(RESULTS / f"optuna_results_{task}.xlsx")
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["Feature", "Best_LR", "Best_Dropout", "Best_LSTM_Hidden", "Best_WD", "Best_Epoch",
                "Test_Pre", "Test_Rec", "Test_AUC", "Test_Acc"])
    ws.append([
        feature,
        best_trial.params["lr"],
        best_trial.params["dropout"],
        best_trial.params["lstm_hidden"],
        best_trial.params["weight_decay"],
        best_trial.user_attrs["best_epoch"],
        round(t_pre, 4), round(t_rec, 4), round(t_auc, 4), round(t_acc, 4),
    ])
    wb.save(results_path)

    return str(save_path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Train BiLSTM+Attention with Optuna")
    parser.add_argument("--task", default="amp", choices=["amp", "aep", "hp"])
    parser.add_argument("--feature", default="pc7")
    parser.add_argument("--trials", type=int, default=30)
    args = parser.parse_args()
    train_attention(args.task, args.feature, args.trials)
