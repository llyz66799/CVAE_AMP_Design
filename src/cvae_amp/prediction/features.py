"""Physicochemical descriptor encoding for peptide sequences.

StandardScaler is fit once on the 20 standard amino acids at init time,
preventing data leakage between train/val/test splits.
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import openpyxl
from sklearn.preprocessing import StandardScaler

from cvae_amp.config.defaults import AA_TO_IDX, MAX_LEN

PAD_IDX = AA_TO_IDX["B"]


class FeatureEncoder:
    """Encodes amino acid sequences into fixed-length feature vectors.

    Each amino acid is mapped to a vector of physicochemical descriptors
    (e.g. hydrophobicity, volume, charge). The StandardScaler is fit once
    on the 20 standard amino acids so feature scales are stable across
    repeated calls to ``encode_file``.
    """

    def __init__(self, feature_path: str) -> None:
        df = pd.read_excel(feature_path, sheet_name=0)
        raw = df.iloc[:20, 1:-1].values.astype(np.float64)

        self.scaler = StandardScaler()
        self.scaler.fit(raw)
        scaled = self.scaler.transform(raw)

        self._id_to_vec: dict[int, np.ndarray] = {}
        for i in range(20):
            self._id_to_vec[i] = scaled[i]
        self._id_to_vec[PAD_IDX] = np.zeros(scaled.shape[1], dtype=np.float64)

        self._dim = scaled.shape[1]

    @property
    def dim(self) -> int:
        return self._dim

    def encode_file(self, path: str, seq_col: str | int | None = None) -> np.ndarray:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        n_rows = ws.max_row - 1

        # Resolve sequence column: by name, by index, or auto-detect
        if isinstance(seq_col, int):
            col_idx = seq_col
        elif isinstance(seq_col, str):
            col_idx = _find_col_by_header(ws, seq_col)
            if col_idx is None:
                raise ValueError(f"Column '{seq_col}' not found in {path}")
        else:
            col_idx = _find_col_by_header(ws, "seq") or 2

        result = np.zeros((n_rows, MAX_LEN, self._dim), dtype=np.float64)

        for i in range(2, ws.max_row + 1):
            idx = i - 2
            seq_val = ws.cell(row=i, column=col_idx).value
            if seq_val is None:
                continue
            seq = str(seq_val).strip().upper()
            padded = seq.ljust(MAX_LEN, "B")[:MAX_LEN]
            for j, aa in enumerate(padded):
                aa_id = AA_TO_IDX.get(aa, PAD_IDX)
                result[idx, j] = self._id_to_vec[aa_id]

        return result


def _find_col_by_header(ws, name: str) -> int | None:
    """Return 1-based column index for header *name*, or None if not found."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == name:
            return col
    return None
